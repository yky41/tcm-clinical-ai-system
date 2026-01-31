import requests
from lxml import html
from openpyxl import Workbook
import re


class zhongyao():
    def __init__(self):
        self.text_all = dict()
        self.url = "http://www.a-hospital.com/w/%E4%B8%AD%E8%8D%AF%E6%96%B9%E5%89%82"
        self.headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/78.0.3904.97 Safari/537.36"}

    def get_parse_html(self, url):
        res = requests.get(url, headers=self.headers)
        res.encoding = "utf-8"
        html_text = res.text
        parse_html = html.etree.HTML(html_text)
        return parse_html

    def get_url(self):
        res = requests.get(self.url, headers=self.headers)
        res.encoding = "utf-8"
        en = html.etree.HTML(res.text)
        text_urls = []

        text_url = en.xpath(f'//*[@id="bodyContent"]/ul[1]/li//a[1]/@href')  # 剂的地址
        if text_url:
            for url in text_url:  # 迭代链接列表
                url1 = "http://www.a-hospital.com" + url
                parse_html_text1 = self.get_parse_html(url1)  # 进入到第一个链接
                text_title = parse_html_text1.xpath('string(//h1[@id="firstHeading"]/text())')

                text_url2 = parse_html_text1.xpath('//*[@class="wikitable"]/tr/td[1]//a/@href')
                for url2 in text_url2:
                    url2 = "http://www.a-hospital.com" + url2
                    text_urls.append((url2,text_title))
        self.get_nate(text_urls)

    def get_text_alias(self, parse_html_text):
        try:
            alias_data = ""
            # 先在 <p> 标签中搜索特定内容
            text_alias_p = parse_html_text.xpath(
                f'//*[@id="bodyContent"]/p[contains(.,"组成") or contains(.,"组 成") or contains(.,"成 分") or contains(.,"成分") or contains(.,"处方")]')
            if text_alias_p:
                # 获取第一个匹配的段落文本
                alias_data = text_alias_p[0].xpath('string()').strip()
            else:
                # 如果没有找到符合条件的 <p> 标签，则直接搜索 <h> 标签
                text_heading = parse_html_text.xpath(
                    f'//*[@id="bodyContent"]/h[contains(.,"组成") or contains(.,"组 成") or contains(.,"成 分") or contains(.,"成分") or contains(.,"处方")]')
                if text_heading:
                    # 获取该 <h> 标签的所有兄弟节点
                    siblings = text_heading[0].xpath('./following-sibling::*')
                    # 寻找兄弟节点中的第一个 <p> 标签
                    for sibling in siblings:
                        if sibling.tag == 'p':
                            alias_data = sibling.xpath('string()').strip()
                            break

        except Exception as e:
            print(e)
            return ""

        return alias_data

    def get_text_smell(self, parse_html_text):
        try:
            smell_data = ""
            text_smell = parse_html_text.xpath(f'//*[@id="bodyContent"]/p[contains(.,"主治")] ')
            if text_smell:
                smell_data = text_smell[0].xpath('string()').strip()

        except:
            return ""
        return smell_data

    def get_text_cure(self, parse_html_text):
        try:
            cure_data = ""
            text_cure = parse_html_text.xpath(f'//*[@id="bodyContent"]/p[contains(.,"用法")] ')
            if text_cure:
                cure_data = text_cure[0].xpath('string()').strip()



            cure_data = cure_data.strip()
        except:
            return ""
        return cure_data

    def get_text_gx(self, parse_html_text):
        try:
            gx_data = ""
            # 获取当前行第一个td标签的文本内容，并拼接到alias_data中
            text_name = parse_html_text.xpath('string(//h1[@id="firstHeading"]/text())')
            if text_name:
                gx_data += ''.join(text_name) # 拼接文本内容，并添加换行符
        # 去除最后一个换行符
            gx_data = gx_data
        except:
            return ""
        return gx_data


    # def get_text_zz(self, parse_html_text):
    #     try:
    #         zz_data = ""
    #
    #         for i in range(1, 15):
    #             # 获取当前行第一个td标签的文本内容，并拼接到alias_data中
    #             text_alias = parse_html_text.xpath(f'//*[@class="wikitable"]/tr[{i}]/td[5]//text()')
    #             if text_alias:
    #                 zz_data += ''.join(text_alias)  # 拼接文本内容，并添加换行符
    #         # 去除最后一个换行符
    #         zz_data = zz_data
    #     except:
    #         return ""
    #     return zz_data

    def get_nate(self, text_urls):
        count = 0
        rows = []
        for link,title in text_urls:
            t_url = link
            parse_html_text = self.get_parse_html(t_url)
            #text_name = parse_html_text.xpath('//*[@id="firstHeading"]/text()')
            text_name=title
            text_name = ''.join(text_name)
            text_alias = self.get_text_alias(parse_html_text)

            text_smell = self.get_text_smell(parse_html_text)
            text_cure = self.get_text_cure(parse_html_text)
            text_gx= self.get_text_gx(parse_html_text)
            #text_zz=self.get_text_zz(parse_html_text)

            wb = Workbook()
            ws = wb.active
            ws['A1'] = '方剂类型'
            ws['B1'] = '方名'
            ws['C1'] = '组成'
            ws['D1'] = '主治'
            ws['E1'] = '用法'


            row = [text_name, text_gx,text_alias ,text_smell, text_cure]

            rows.append(row)
            for new_row in rows:
                ws.append(new_row)
            count += 1
            print(count)
            wb.save('部.xlsx')#全都保存在了这里


zhongyao = zhongyao()
zhongyao.get_url()
